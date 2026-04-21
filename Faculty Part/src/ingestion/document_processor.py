"""
Document processing for PDFs, images, and text files.

PDF extraction uses PyMuPDF (fitz) for correct reading order and table detection.
Falls back to pypdf if fitz is unavailable.
"""

import io
import json
import logging
import re
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Any, Dict, List, Optional

import pytesseract
from PIL import Image, ImageOps

logger = logging.getLogger(__name__)

# P1: Try PyMuPDF first, fall back to pypdf
try:
    import fitz  # PyMuPDF
    _HAS_FITZ = True
except ImportError:
    _HAS_FITZ = False
    from pypdf import PdfReader
    logger.warning("PyMuPDF not installed — falling back to pypdf. "
                   "Install: pip install pymupdf  for better extraction quality.")


def _normalize_text(text: str) -> str:
    """
    P2: Unicode normalization + ligature/hyphen handling.

    Fixes:
    - Ligatures: ﬁ → fi, ﬂ → fl, ﬀ → ff
    - Non-breaking hyphens → ASCII hyphen (critical for form codes like HR-LA-01)
    - Smart quotes → ASCII
    - P3: Hyphenation at line break: appli-\ncation → application
    """
    # NFKC decomposes ligatures and normalizes compatibility characters
    text = unicodedata.normalize("NFKC", text)

    # Non-breaking and soft hyphens → ASCII hyphen
    text = text.replace("\u2010", "-").replace("\u2011", "-").replace("\u00ad", "")

    # Smart quotes
    text = text.replace("\u2019", "'").replace("\u2018", "'")
    text = text.replace("\u201c", '"').replace("\u201d", '"')

    # Em/en dash → hyphen
    text = text.replace("\u2014", "-").replace("\u2013", "-")

    # P3: Merge hyphenated line breaks (lowercase-hyphen-newline-lowercase only)
    text = re.sub(r'([a-z])-\n([a-z])', r'\1\2', text)

    return text


def _ocr_looks_valid(text: str) -> bool:
    """P6: Gibberish check for OCR output."""
    if len(text.strip()) < 20:
        return False
    alpha_ratio = sum(c.isalpha() for c in text) / len(text)
    if alpha_ratio < 0.5:
        return False
    words = text.split()
    valid_words = sum(1 for w in words if 3 <= len(w) <= 15)
    return valid_words / max(len(words), 1) >= 0.5


class DocumentProcessor:
    """
    Process various document types into text.

    Handles:
    - PDFs (PyMuPDF preferred, pypdf fallback, OCR for scanned pages)
    - Images (OCR + visual description)
    - Text files, JSON, CSV, Excel
    """

    def __init__(self, tesseract_path: Optional[str] = None):
        if tesseract_path:
            pytesseract.pytesseract.tesseract_cmd = tesseract_path

    def process_document(
        self,
        file_path: Path,
        doc_metadata: Dict[str, Any]
    ) -> Dict[str, Any]:
        suffix = file_path.suffix.lower()

        if suffix == '.pdf':
            return self._process_pdf(file_path, doc_metadata)
        elif suffix in ['.png', '.jpg', '.jpeg', '.tiff']:
            return self._process_image(file_path, doc_metadata)
        elif suffix in ['.txt', '.md']:
            return self._process_text(file_path, doc_metadata)
        elif suffix == '.json':
            return self._process_json(file_path, doc_metadata)
        elif suffix == '.csv':
            return self._process_csv(file_path, doc_metadata)
        elif suffix in ['.xlsx', '.xls']:
            return self._process_excel(file_path, doc_metadata)
        else:
            raise ValueError(f"Unsupported file type: {suffix}")

    # ── PDF ──────────────────────────────────────────────────────────────────

    def _process_pdf(
        self,
        file_path: Path,
        doc_metadata: Dict[str, Any]
    ) -> Dict[str, Any]:
        if _HAS_FITZ:
            return self._process_pdf_fitz(file_path, doc_metadata)
        return self._process_pdf_pypdf(file_path, doc_metadata)

    def _process_pdf_fitz(
        self,
        file_path: Path,
        doc_metadata: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        P1: PyMuPDF extraction with correct reading order.

        Sorts text blocks by (y, x) for multi-column layouts.
        P5: Detects and serializes tables as pipe-delimited rows.
        P4: Auto-detects repeated header/footer lines and strips them.
        """
        doc = fitz.open(str(file_path))
        pages_content = []
        ocr_pages = 0
        ocr_failed_pages = 0
        empty_pages = []
        tables_found = 0

        for page_num, page in enumerate(doc):
            # P5: Extract tables first
            table_texts = []
            try:
                for tbl in page.find_tables():
                    rows = tbl.extract()
                    if rows:
                        table_texts.append(
                            "\n".join(
                                " | ".join(str(c) if c else "" for c in row)
                                for row in rows
                            )
                        )
                        tables_found += 1
            except Exception:
                pass

            # P1: Block-level text with reading-order sort
            blocks = page.get_text("blocks")
            blocks.sort(key=lambda b: (round(b[1] / 10), b[0]))
            text = "\n".join(b[4] for b in blocks if b[6] == 0)  # type 0 = text

            # Append table text after prose
            if table_texts:
                text = text + "\n\n" + "\n\n".join(table_texts)

            # P2: normalize
            text = _normalize_text(text)

            if len(text.strip()) < 50:
                ocr_text = self._ocr_pdf_page(file_path, page_num + 1)
                if ocr_text and _ocr_looks_valid(ocr_text):
                    text = ocr_text
                    ocr_pages += 1
                else:
                    ocr_failed_pages += 1
                    if len(text.strip()) < 20:
                        empty_pages.append(page_num + 1)

            text = self._join_continuation_lines(text)
            pages_content.append({"page_num": page_num + 1, "content": text})

        doc.close()

        # P4: Auto-detect and strip repeated header/footer lines
        pages_content, header_footer_lines = self._strip_repeated_lines(pages_content)

        full_text = "\n\n".join(p["content"] for p in pages_content)
        page_count = len(pages_content)

        self._check_extraction_quality(full_text, page_count, file_path, ocr_pages, ocr_failed_pages, empty_pages)

        # P10: structured extraction log
        self._log_extraction({
            "file": file_path.name,
            "extractor": "fitz",
            "page_count": page_count,
            "total_chars": len(full_text.strip()),
            "ocr_pages": ocr_pages,
            "ocr_failed": ocr_failed_pages,
            "empty_pages": empty_pages,
            "tables_detected": tables_found,
            "header_footer_lines_stripped": len(header_footer_lines),
        })

        return {
            "content": full_text,
            "pages": pages_content,
            "images": [],
            "metadata": {
                **doc_metadata,
                "page_count": page_count,
                "has_images": False,
                "ocr_pages": ocr_pages,
                "ocr_failed_pages": ocr_failed_pages,
            }
        }

    def _process_pdf_pypdf(
        self,
        file_path: Path,
        doc_metadata: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Fallback PDF extraction using pypdf."""
        reader = PdfReader(file_path)
        pages_content = []
        images_found = []
        ocr_pages = 0
        ocr_failed_pages = 0
        empty_pages = []

        for page_num, page in enumerate(reader.pages):
            try:
                text = page.extract_text() or ""
            except Exception as e:
                logger.warning("pypdf extract_text failed on page %d: %s", page_num + 1, e)
                text = ""

            text = _normalize_text(text)

            if len(text.strip()) < 50:
                ocr_text = self._ocr_pdf_page(file_path, page_num + 1)
                if ocr_text and _ocr_looks_valid(ocr_text):
                    text = ocr_text
                    ocr_pages += 1
                else:
                    ocr_failed_pages += 1
                    if len(text.strip()) < 20:
                        empty_pages.append(page_num + 1)

            text = self._join_continuation_lines(text)
            pages_content.append({"page_num": page_num + 1, "content": text})

            try:
                if '/XObject' in page['/Resources']:
                    xobjects = page['/Resources']['/XObject'].get_object()
                    for obj_name in xobjects:
                        obj = xobjects[obj_name]
                        if obj['/Subtype'] == '/Image':
                            image_data = self._extract_pdf_image(obj)
                            if image_data:
                                images_found.append({
                                    "page_num": page_num + 1,
                                    "image_name": obj_name,
                                    **image_data,
                                })
            except Exception:
                pass

        pages_content, header_footer_lines = self._strip_repeated_lines(pages_content)
        full_text = "\n\n".join(p["content"] for p in pages_content)
        page_count = len(reader.pages)

        self._check_extraction_quality(full_text, page_count, file_path, ocr_pages, ocr_failed_pages, empty_pages)

        self._log_extraction({
            "file": file_path.name,
            "extractor": "pypdf",
            "page_count": page_count,
            "total_chars": len(full_text.strip()),
            "ocr_pages": ocr_pages,
            "ocr_failed": ocr_failed_pages,
            "empty_pages": empty_pages,
            "tables_detected": 0,
            "header_footer_lines_stripped": len(header_footer_lines),
        })

        return {
            "content": full_text,
            "pages": pages_content,
            "images": images_found,
            "metadata": {
                **doc_metadata,
                "page_count": page_count,
                "has_images": len(images_found) > 0,
                "ocr_pages": ocr_pages,
                "ocr_failed_pages": ocr_failed_pages,
            }
        }

    def _strip_repeated_lines(self, pages_content: List[Dict]) -> tuple:
        """
        P4: Detect lines appearing on >60% of pages and strip them.

        Returns (updated_pages_content, set_of_stripped_lines).
        """
        page_count = len(pages_content)
        if page_count < 3:
            return pages_content, set()

        line_freq: Counter = Counter()
        for page in pages_content:
            for line in set(page["content"].split("\n")):
                stripped = line.strip()
                if stripped:
                    line_freq[stripped] += 1

        threshold = 0.6
        header_footer = {
            line for line, count in line_freq.items()
            if count / page_count >= threshold
        }

        if not header_footer:
            return pages_content, set()

        for page in pages_content:
            page["content"] = "\n".join(
                line for line in page["content"].split("\n")
                if line.strip() not in header_footer
            )

        return pages_content, header_footer

    def _check_extraction_quality(
        self,
        full_text: str,
        page_count: int,
        file_path: Path,
        ocr_pages: int,
        ocr_failed_pages: int,
        empty_pages: List[int],
    ) -> None:
        """Raise RuntimeError if extraction quality is too low."""
        total_chars = len(full_text.strip())
        min_expected = max(200, page_count * 50)
        if total_chars < min_expected:
            raise RuntimeError(
                f"PDF extraction produced only {total_chars} chars across "
                f"{page_count} pages for '{file_path.name}' "
                f"(ocr_pages={ocr_pages}, ocr_failed_pages={ocr_failed_pages}). "
                f"Repair with 'mutool clean -gggg in.pdf out.pdf' then retry."
            )

        if empty_pages:
            threshold = max(1, int(page_count * 0.2))
            if len(empty_pages) > threshold:
                raise RuntimeError(
                    f"'{file_path.name}': {len(empty_pages)} pages produced <20 chars "
                    f"and OCR also failed (pages: {empty_pages}). "
                    f"Repair with 'mutool clean -gggg in.pdf out.pdf' then retry."
                )

    # ── Image ────────────────────────────────────────────────────────────────

    def _process_image(
        self,
        file_path: Path,
        doc_metadata: Dict[str, Any]
    ) -> Dict[str, Any]:
        image = Image.open(file_path)
        # P9: basic preprocessing for better OCR
        image = ImageOps.grayscale(image)
        image = ImageOps.autocontrast(image)
        ocr_text = pytesseract.image_to_string(image)
        description = self._generate_image_description(image)
        return {
            "content": ocr_text,
            "visual_description": description,
            "metadata": {
                **doc_metadata,
                "image_size": image.size,
                "image_mode": image.mode,
                "has_text": len(ocr_text.strip()) > 0,
            }
        }

    # ── Text ─────────────────────────────────────────────────────────────────

    def _process_text(
        self,
        file_path: Path,
        doc_metadata: Dict[str, Any]
    ) -> Dict[str, Any]:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return {"content": content, "metadata": doc_metadata}

    # ── JSON ─────────────────────────────────────────────────────────────────

    def _process_json(
        self,
        file_path: Path,
        doc_metadata: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        Process JSON faculty data.

        Returns one clean profile block per faculty member separated by
        '=' * 60 so DocumentChunker._chunk_faculty_profile can split them.
        """
        content_parts = []

        with open(file_path, 'r', encoding='utf-8-sig') as f:
            content = f.read().strip()

        try:
            data = json.loads(content)
        except json.JSONDecodeError:
            data = None

        if isinstance(data, list):
            entries = data
        elif isinstance(data, dict):
            entries = [data]
        else:
            # JSONL fallback
            entries = []
            for line in content.splitlines():
                line = line.strip()
                if not line:
                    continue
                try:
                    entries.append(json.loads(line))
                except json.JSONDecodeError:
                    pass

        if not entries:
            raise ValueError(f"No valid JSON entries found in {file_path.name}")

        for i, obj in enumerate(entries, 1):
            if not isinstance(obj, dict):
                continue
            text = obj.get('text', '') or self._build_faculty_text_from_dict(obj)
            name = obj.get('name', f'Entry {i}')
            profile_url = obj.get('profile_url', '')
            # P8: return clean profile block; chunker decides where to split
            block = f"Faculty: {name}\n"
            if profile_url:
                block += f"Profile: {profile_url}\n"
            block += "\n" + text.strip()
            content_parts.append(block)

        separator = "\n" + "=" * 60 + "\n\n"
        return {
            "content": separator.join(content_parts),
            "metadata": {**doc_metadata, "entry_count": len(content_parts), "format": "json"}
        }

    def _build_faculty_text_from_dict(self, obj: Dict[str, Any]) -> str:
        parts = []
        if 'name' in obj:
            parts.append(f"Name: {obj['name']}")
        for field in ['qualification', 'experience', 'research_interests']:
            if field in obj:
                value = str(obj[field]).strip()
                if value and value.lower() != 'not specified':
                    parts.append(f"{field.replace('_', ' ').title()}: {value}")
        if 'publications' in obj:
            pubs = obj['publications']
            if isinstance(pubs, dict):
                for pub_type, pub_list in pubs.items():
                    if isinstance(pub_list, list) and pub_list:
                        parts.append(f"\n{pub_type.title()}:")
                        for pub in pub_list:
                            parts.append(f"  - {pub}")
            elif isinstance(pubs, list) and pubs:
                parts.append("Publications:")
                for pub in pubs:
                    parts.append(f"  - {pub}")
        if 'awards' in obj:
            awards = obj['awards']
            if isinstance(awards, list) and awards:
                parts.append("\nAwards:")
                for award in awards:
                    parts.append(f"  - {award}")
        return "\n".join(parts)

    # ── CSV ──────────────────────────────────────────────────────────────────

    def _process_csv(
        self,
        file_path: Path,
        doc_metadata: Dict[str, Any]
    ) -> Dict[str, Any]:
        import csv
        encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252', 'iso-8859-1']
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    reader = csv.DictReader(f)
                    headers = reader.fieldnames
                    rows = list(reader)
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
        else:
            raise ValueError(f"Could not decode {file_path.name} with any common encoding")

        title = doc_metadata.get('title', file_path.stem)
        content_parts = [f"{title}\n", "=" * len(title) + "\n\n"]
        for i, row in enumerate(rows, 1):
            content_parts.append(f"Entry {i}:\n")
            for key, value in row.items():
                if value:
                    content_parts.append(f"  {key}: {value}\n")
            content_parts.append("\n")

        return {
            "content": "".join(content_parts),
            "metadata": {**doc_metadata, "row_count": len(rows), "columns": list(headers) if headers else []},
        }

    # ── Excel ────────────────────────────────────────────────────────────────

    def _process_excel(
        self,
        file_path: Path,
        doc_metadata: Dict[str, Any]
    ) -> Dict[str, Any]:
        """P7: Process all sheets, not just the active one."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl required. Install: pip install openpyxl")

        workbook = openpyxl.load_workbook(file_path, data_only=True)
        title = doc_metadata.get('title', file_path.stem)
        content_parts = [f"{title}\n", "=" * len(title) + "\n\n"]
        total_rows = 0

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            content_parts.append(f"\n## Sheet: {sheet_name}\n\n")
            headers = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    row_dict = dict(zip(headers, row))
                    content_parts.append(f"Entry {total_rows + 1}:\n")
                    for key, value in row_dict.items():
                        if value is not None and str(value).strip():
                            content_parts.append(f"  {key}: {value}\n")
                    content_parts.append("\n")
                    total_rows += 1

        return {
            "content": "".join(content_parts),
            "metadata": {
                **doc_metadata,
                "row_count": total_rows,
                "sheet_names": workbook.sheetnames,
            }
        }

    # ── OCR helpers ──────────────────────────────────────────────────────────

    def _ocr_pdf_page(self, file_path: Path, page_num: int) -> str:
        try:
            from pdf2image import convert_from_path
        except ImportError:
            logger.warning("[OCR] pdf2image not installed. Install: pip install pdf2image")
            return ""

        try:
            images = convert_from_path(str(file_path), first_page=page_num, last_page=page_num, dpi=300)
        except Exception as e:
            logger.warning("[OCR] pdf2image failed for %s p%d: %s", file_path.name, page_num, e)
            return ""

        if not images:
            return ""

        text = ""
        try:
            # P9: preprocess for better OCR accuracy
            img = ImageOps.grayscale(images[0])
            img = ImageOps.autocontrast(img)
            text = pytesseract.image_to_string(img)
        except Exception as e:
            logger.warning("[OCR] tesseract failed for %s p%d: %s", file_path.name, page_num, e)
        finally:
            for img in images:
                img.close()

        return text or ""

    def _extract_pdf_image(self, image_obj) -> Optional[Dict[str, str]]:
        try:
            data = image_obj.get_data()
            image = Image.open(io.BytesIO(data))
            ocr_text = pytesseract.image_to_string(image)
            description = self._generate_image_description(image)
            return {"ocr_text": ocr_text, "description": description}
        except Exception as e:
            logger.debug("Error extracting image: %s", e)
            return None

    def _generate_image_description(self, image: Image.Image) -> str:
        return f"[Image: {image.size[0]}x{image.size[1]} pixels]"

    # ── Text cleaning ────────────────────────────────────────────────────────

    def _join_continuation_lines(self, text: str) -> str:
        """Join continuation lines while preserving table rows."""
        _table_pat = re.compile(r'^\w[^\n]{1,40}\s{3,}.+')
        lines = text.split('\n')
        cleaned = []

        for line in lines:
            stripped = line.strip()
            if not stripped:
                cleaned.append('')
                continue

            prev = cleaned[-1] if cleaned else ''
            prev_stripped = prev.strip()

            is_table_line = (
                '   ' in stripped or
                line.startswith('   ') or
                (prev_stripped and '   ' in prev_stripped) or
                _table_pat.match(stripped) or
                (prev_stripped and _table_pat.match(prev_stripped))
            )

            if (not is_table_line and prev_stripped and stripped and
                    (stripped[0].islower() or
                     stripped.startswith('(') or
                     re.match(r'^(with|or|and|for|by|of|in|at|to)\s', stripped, re.IGNORECASE))):
                cleaned[-1] = cleaned[-1].rstrip() + ' ' + stripped
            else:
                cleaned.append(line)

        return '\n'.join(cleaned)

    # ── Extraction logging ───────────────────────────────────────────────────

    def _log_extraction(self, stats: Dict[str, Any]) -> None:
        """P10: Write per-document extraction stats to JSONL log."""
        import json as _json
        from datetime import datetime
        log_path = Path("logs/extraction.jsonl")
        try:
            log_path.parent.mkdir(parents=True, exist_ok=True)
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(_json.dumps({"timestamp": datetime.utcnow().isoformat(), **stats}) + "\n")
        except Exception as e:
            logger.debug("Failed to write extraction log: %s", e)
